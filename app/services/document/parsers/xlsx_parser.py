import logging
from openpyxl import load_workbook
from .base_parser import BaseParser, ParseResult

logger = logging.getLogger(__name__)


class XlsxParser(BaseParser):
    """Excel 文档解析器，使用 openpyxl"""

    def parse(self, file_path: str) -> ParseResult:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheets_text = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_parts = [f'--- Sheet: {sheet_name} ---']
            for row in ws.iter_rows(values_only=True):
                row_text = [str(cell) if cell is not None else '' for cell in row]
                # 跳过全空行
                if any(cell.strip() for cell in row_text):
                    sheet_parts.append(' | '.join(row_text))

            if len(sheet_parts) > 1:
                sheets_text.append('\n'.join(sheet_parts))

        wb.close()

        content = '\n\n'.join(sheets_text)

        import os
        title = os.path.splitext(os.path.basename(file_path))[0]

        return ParseResult(
            content=content.strip(),
            title=title,
            metadata={
                'sheet_count': len(wb.sheetnames) if hasattr(wb, 'sheetnames') else 0,
                'format': 'xlsx',
            }
        )

    def supported_extensions(self) -> list:
        return ['.xlsx', '.xls']
